# app.py — 영단어 학습 & 시험 앱 (Streamlit) — 모바일 최적화 + 단어 관리

import re
import io
import os
import json
import random
from datetime import datetime

import base64

import openpyxl
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from streamlit_js_eval import streamlit_js_eval

# ============================================================
# 페이지 설정 (반드시 첫 번째 Streamlit 호출)
# ============================================================
st.set_page_config(
    page_title="영단어 학습기",
    page_icon="📚",
    layout="centered",
    initial_sidebar_state="collapsed",
)

# ============================================================
# 전역 CSS — 모바일 최적화
# ============================================================
st.markdown("""
<style>
/* ── 전체 패딩 축소 (모바일) ── */
.block-container { padding-top: 1rem !important; padding-bottom: 2rem !important; }

/* ── 메트릭 카드 ── */
[data-testid="metric-container"] {
    background: #f0f4ff;
    border: 1px solid #c5cae9;
    border-radius: 10px;
    padding: 10px 12px;
}

/* ── 버튼 터치 타깃 크게 ── */
.stButton > button {
    min-height: 52px !important;
    font-size: 1rem !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
}

/* ── 플래시카드 앞면 ── */
.card-front {
    text-align: center;
    font-size: clamp(1.8rem, 6vw, 3rem);
    font-weight: 800;
    letter-spacing: 2px;
    color: #ffffff;
    padding: 40px 20px 32px;
    line-height: 1.2;
}

/* ── 플래시카드 뒷면 ── */
.card-back {
    text-align: center;
    font-size: clamp(1.4rem, 4vw, 2rem);
    font-weight: 700;
    color: #ffffff;
    padding: 16px 20px 24px;
    line-height: 1.4;
}

/* ── 무한 시험 문제 텍스트 ── */
.inf-question {
    text-align: center;
    font-size: clamp(1.6rem, 5vw, 2.6rem);
    font-weight: 800;
    letter-spacing: 1px;
    color: #1a237e;
    padding: 28px 16px 20px;
    line-height: 1.3;
}

/* ── 토익 빈칸 문장 ── */
.toeic-sentence {
    font-size: clamp(1rem, 3vw, 1.25rem);
    line-height: 1.9;
    padding: 18px 20px;
    border-left: 5px solid #1565C0;
    background: #e8f0fe;
    border-radius: 6px;
    font-style: italic;
    color: #1a237e;
    margin: 8px 0 16px;
}

/* ── 시험 문제 번호 ── */
.q-label { font-size: 1.05rem; font-weight: 600; margin-bottom: 4px; }
.q-num {
    display: inline-block;
    background: #3949ab; color: white;
    border-radius: 6px; padding: 1px 9px;
    font-size: 0.95rem; margin-right: 6px;
}

/* ── 모바일: 컬럼이 너무 좁아지지 않도록 ── */
@media (max-width: 640px) {
    [data-testid="column"] { min-width: 40% !important; }
    .card-front  { padding: 28px 12px 20px; }
    .stTabs [data-baseweb="tab"] { font-size: 0.78rem !important; padding: 6px 8px !important; }
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# 경로
# ============================================================
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH    = os.path.join(BASE_DIR, "Word.xlsx")
PROGRESS_PATH = os.path.join(BASE_DIR, "progress.json")

SHEET_NAME = "영어 단어장"

# ============================================================
# 데이터 로드 / 저장
# ============================================================
@st.cache_data
def load_words():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]
    result = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[1]:
            continue
        result.append({
            "no":       row[0],
            "word":     str(row[1]).strip(),
            "meaning":  str(row[2]).strip() if row[2] else "",
            "example":  str(row[3]).strip() if row[3] else "",
            "category": str(row[8]).strip() if len(row) > 8 and row[8] else "기타",
            "_extra":   list(row[4:8]),   # 열 5~8 보존
        })
    return result


def save_words_to_excel(word_list):
    """word_list를 Excel 파일에 저장하고 캐시를 초기화."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    # 기존 데이터 행 지우기
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 다시 쓰기
    for i, w in enumerate(word_list, start=2):
        ws.cell(row=i, column=1, value=w["no"])
        ws.cell(row=i, column=2, value=w["word"])
        ws.cell(row=i, column=3, value=w["meaning"])
        ws.cell(row=i, column=4, value=w["example"])
        for j, val in enumerate(w.get("_extra", [None] * 4), start=5):
            ws.cell(row=i, column=j, value=val)
        ws.cell(row=i, column=9, value=w["category"])

    wb.save(EXCEL_PATH)
    load_words.clear()   # 캐시 무효화


LS_KEY = "ew_progress"   # localStorage 키


def load_progress():
    if os.path.exists(PROGRESS_PATH):
        with open(PROGRESS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"memorized": [], "wrong": [], "correct": [], "sessions": []}


def save_progress():
    data = {
        "memorized": list(st.session_state.memorized),
        "wrong":     list(st.session_state.wrong),
        "correct":   list(st.session_state.correct_words),
        "sessions":  st.session_state.sessions,
    }
    # 로컬 파일 저장 (PC 로컬 실행 시)
    try:
        with open(PROGRESS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    # 브라우저 localStorage 저장 (클라우드/모바일 지속성)
    _save_to_localstorage(data)


def _save_to_localstorage(data: dict):
    """base64로 인코딩해 localStorage에 저장 (특수문자 이슈 방지)."""
    raw  = json.dumps(data, ensure_ascii=False)
    b64  = base64.b64encode(raw.encode()).decode()
    components.html(f"""
    <script>
    (function() {{
        try {{
            localStorage.setItem('{LS_KEY}', atob('{b64}'));
        }} catch(e) {{ console.warn('localStorage write failed', e); }}
    }})();
    </script>
    """, height=0, width=0)


# ============================================================
# 헬퍼 함수
# ============================================================
def filter_words(word_list, cats, scope):
    pool = [w for w in word_list if w["category"] in cats]
    if scope == "미암기만":
        pool = [w for w in pool if w["no"] not in st.session_state.memorized]
    elif scope == "오답만":
        pool = [w for w in pool if w["no"] in st.session_state.wrong]
    elif scope == "정답 맞춘":
        pool = [w for w in pool if w["no"] in st.session_state.correct_words]
    return pool


def make_choices(q, direction, all_words):
    if direction == "영어 → 뜻":
        correct = q["meaning"]
        pool    = [w["meaning"] for w in all_words if w["no"] != q["no"] and w["meaning"]]
    else:
        correct = q["word"]
        pool    = [w["word"] for w in all_words if w["no"] != q["no"]]
    choices = random.sample(pool, min(3, len(pool))) + [correct]
    random.shuffle(choices)
    return choices, correct


def make_toeic_question(q, all_words):
    if not q["example"]:
        return None
    pattern = re.compile(re.escape(q["word"]), re.IGNORECASE)
    blanked = pattern.sub("_____ ", q["example"], count=1)
    if blanked == q["example"]:
        return None
    wrong_pool = [w["word"] for w in all_words if w["no"] != q["no"]]
    choices    = random.sample(wrong_pool, min(3, len(wrong_pool))) + [q["word"]]
    random.shuffle(choices)
    return {"sentence": blanked.strip(), "choices": choices, "correct": q["word"]}


def mark_correct(no):
    st.session_state.correct_words.add(no)
    st.session_state.wrong.discard(no)


def mark_wrong(no):
    st.session_state.wrong.add(no)
    st.session_state.correct_words.discard(no)


def tts_button(word: str, key: str = "tts"):
    safe = word.replace("'", "\\'").replace('"', '\\"')
    components.html(f"""
    <button onclick="speak()" title="{safe} 발음 듣기" style="
        background:#1565C0;color:white;border:none;border-radius:10px;
        padding:10px 20px;font-size:1.1rem;cursor:pointer;
        display:flex;align-items:center;gap:6px;margin:4px auto;
        min-height:48px;touch-action:manipulation;">
        🔊 발음 듣기
    </button>
    <script>
    function speak() {{
        window.speechSynthesis.cancel();
        var u = new SpeechSynthesisUtterance("{safe}");
        u.lang = "en-US";
        u.rate = 0.85;
        window.speechSynthesis.speak(u);
    }}
    </script>
    """, height=52)


def make_excel_bytes(word_list, extra_cols=None):
    rows = []
    for w in word_list:
        row = {"번호": w["no"], "단어": w["word"], "뜻": w["meaning"],
               "예문": w["example"], "카테고리": w["category"]}
        if extra_cols:
            row.update(extra_cols.get(w["no"], {}))
        rows.append(row)
    buf = io.BytesIO()
    pd.DataFrame(rows).to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf


def start_regular_test(pool, n, mode, direction):
    questions = random.sample(pool, n)
    options, toeic_data = {}, {}
    for q in questions:
        if mode == "4지선다":
            opts, _ = make_choices(q, direction, words)
            options[q["no"]] = opts
        elif mode == "토익식 빈칸":
            td = make_toeic_question(q, words)
            if td:
                toeic_data[q["no"]] = td
            else:
                opts, _ = make_choices(q, "영어 → 뜻", words)
                options[q["no"]] = opts
    st.session_state.update({
        "test_questions":  questions,
        "test_options":    options,
        "test_toeic_data": toeic_data,
        "test_mode_val":   mode,
        "test_dir_val":    direction,
        "test_active":     True,
        "test_submitted":  False,
        "test_result":     None,
    })


def pick_next_inf():
    pool      = st.session_state.inf_pool
    mode      = st.session_state.inf_mode
    direction = st.session_state.inf_dir
    if not pool:
        return
    q = random.choice(pool)
    if mode == "토익식 빈칸":
        td = make_toeic_question(q, words)
        if td is None:
            for _ in range(10):
                q  = random.choice(pool)
                td = make_toeic_question(q, words)
                if td:
                    break
        st.session_state.inf_toeic_data  = td
        st.session_state.inf_correct_ans = td["correct"] if td else q["word"]
        choices = td["choices"] if td else []
    else:
        choices, correct = make_choices(q, direction, words)
        st.session_state.inf_toeic_data  = None
        st.session_state.inf_correct_ans = correct
    st.session_state.update({
        "inf_current":  q,
        "inf_choices":  choices,
        "inf_answered": False,
        "inf_selected": "",
        "inf_correct":  False,
    })


# ============================================================
# 세션 초기화
# ============================================================
words          = load_words()
all_categories = sorted(set(w["category"] for w in words))

if "initialized" not in st.session_state:
    # localStorage를 먼저 확인 (클라우드 재시작 후에도 진도 복원)
    ls_raw = streamlit_js_eval(
        js_expressions=f"localStorage.getItem('{LS_KEY}')",
        want_output=True,
        key="ls_init",
    )
    if ls_raw is None:
        # 첫 렌더: JS 결과 대기 중 — 잠시 후 자동으로 재실행됩니다
        st.markdown("⏳ 로딩 중...")
        st.stop()

    # 두 번째 렌더: ls_raw 에 값이 있음
    p: dict = {}
    if ls_raw and ls_raw not in ("null", "undefined", ""):
        try:
            p = json.loads(ls_raw)
        except Exception:
            p = load_progress()
    else:
        p = load_progress()

    st.session_state.memorized     = set(p.get("memorized", []))
    st.session_state.wrong         = set(p.get("wrong", []))
    st.session_state.correct_words = set(p.get("correct", []))
    st.session_state.sessions      = p.get("sessions", [])
    st.session_state.initialized   = True

DEFAULTS = {
    # 플래시카드
    "card_idx": 0, "card_revealed": False,
    "card_list": [], "card_filter_sig": None,
    # 일반 시험
    "test_active": False, "test_submitted": False,
    "test_questions": [], "test_result": None,
    "test_options": {}, "test_toeic_data": {},
    # 무한 시험
    "inf_active": False, "inf_pool": [],
    "inf_dir": "영어 → 뜻", "inf_mode": "4지선다",
    "inf_current": None, "inf_choices": [],
    "inf_correct_ans": "", "inf_toeic_data": None,
    "inf_answered": False, "inf_correct": False,
    "inf_selected": "",
    "inf_session_correct": 0, "inf_session_total": 0,
    "inf_session_wrong_nos": [], "inf_session_correct_nos": [],
    # 단어 관리
    "mgmt_edit_no": None,
    "mgmt_confirm_delete": None,
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ============================================================
# 헤더
# ============================================================
st.title("📚 영단어 학습기")

total     = len(words)
mem_count = len(st.session_state.memorized)
cor_count = len(st.session_state.correct_words)

c1, c2, c3, c4 = st.columns(4)
c1.metric("전체",      f"{total}개")
c2.metric("암기 완료", f"{mem_count}개")
c3.metric("🎯 정답",   f"{cor_count}개")
c4.metric("❌ 오답",   f"{len(st.session_state.wrong)}개")

st.divider()

tab_study, tab_test, tab_inf, tab_wordbook, tab_mgmt, tab_stats = st.tabs(
    ["📖 학습", "📝 시험", "♾️ 무한", "📋 단어장", "✏️ 단어 관리", "📊 통계"]
)


# ============================================================
# TAB 1 — 학습 (플래시카드)
# ============================================================
with tab_study:
    st.subheader("📖 플래시카드 학습")

    study_cats  = st.multiselect("카테고리", all_categories,
                                 default=all_categories, key="study_cats")
    col_dir, col_scope = st.columns(2)
    with col_dir:
        study_dir = st.radio("방향", ["영어 → 뜻", "뜻 → 영어"],
                             horizontal=True, key="study_dir")
    with col_scope:
        study_scope = st.radio("범위", ["전체", "미암기만", "오답만", "정답 맞춘"],
                               horizontal=True, key="study_scope")

    filtered = filter_words(words, study_cats, study_scope)

    if not filtered:
        st.info("학습할 단어가 없어요. 필터를 조정해보세요.")
    else:
        sig = (tuple(sorted(study_cats)), study_scope)
        if st.session_state.card_filter_sig != sig:
            st.session_state.card_list       = filtered.copy()
            random.shuffle(st.session_state.card_list)
            st.session_state.card_idx        = 0
            st.session_state.card_revealed   = False
            st.session_state.card_filter_sig = sig

        card_list = st.session_state.card_list
        idx  = min(st.session_state.card_idx, len(card_list) - 1)
        st.session_state.card_idx = idx
        card = card_list[idx]

        st.progress((idx + 1) / len(card_list), text=f"{idx + 1} / {len(card_list)}")

        is_mem     = card["no"] in st.session_state.memorized
        is_correct = card["no"] in st.session_state.correct_words
        is_wrong   = card["no"] in st.session_state.wrong

        with st.container(border=True):
            front = card["word"] if study_dir == "영어 → 뜻" else card["meaning"]
            st.markdown(f"<div class='card-front'>{front}</div>",
                        unsafe_allow_html=True)
            tts_button(card["word"], key=f"tts_card_{idx}")
            badge = ("🎯 정답 맞춤" if is_correct else "") + \
                    (" ❌ 오답" if is_wrong else "") + \
                    (" ✅ 암기완료" if is_mem else "")
            st.caption(f"카테고리: {card['category']}  No.{card['no']}  {badge}")

            if st.session_state.card_revealed:
                st.divider()
                back = card["meaning"] if study_dir == "영어 → 뜻" else card["word"]
                st.markdown(f"<div class='card-back'>{back}</div>",
                            unsafe_allow_html=True)
                if card["example"]:
                    with st.expander("📖 예문"):
                        st.write(card["example"])

        # 버튼 두 줄 배치 (모바일 친화)
        b1, b2, b3 = st.columns(3)
        with b1:
            if st.button("⬅️ 이전", use_container_width=True):
                st.session_state.card_idx = max(0, idx - 1)
                st.session_state.card_revealed = False
                st.rerun()
        with b2:
            if not st.session_state.card_revealed:
                if st.button("🔍 정답 보기", use_container_width=True, type="primary"):
                    st.session_state.card_revealed = True
                    st.rerun()
            else:
                st.button("🔍 정답 보기", use_container_width=True,
                          type="primary", disabled=True)
        with b3:
            if st.button("➡️ 다음", use_container_width=True):
                st.session_state.card_idx = min(len(card_list) - 1, idx + 1)
                st.session_state.card_revealed = False
                st.rerun()

        b4, b5 = st.columns(2)
        with b4:
            label = "↩️ 암기 취소" if is_mem else "✅ 암기 완료"
            if st.button(label, use_container_width=True):
                if is_mem:
                    st.session_state.memorized.discard(card["no"])
                else:
                    st.session_state.memorized.add(card["no"])
                save_progress()
                st.rerun()
        with b5:
            if st.button("🔀 섞기", use_container_width=True):
                random.shuffle(st.session_state.card_list)
                st.session_state.card_idx      = 0
                st.session_state.card_revealed = False
                st.rerun()


# ============================================================
# TAB 2 — 시험
# ============================================================
with tab_test:

    if not st.session_state.test_active:
        st.subheader("📝 시험 설정")

        t_cats  = st.multiselect("카테고리", all_categories,
                                 default=all_categories, key="t_cats")
        tc1, tc2, tc3 = st.columns(3)
        with tc1:
            t_scope = st.radio("범위", ["전체", "미암기만", "오답만", "정답 맞춘"], key="t_scope")
        with tc2:
            t_mode  = st.radio("유형", ["4지선다", "주관식", "토익식 빈칸"], key="t_mode")
        with tc3:
            t_dir   = st.radio("방향", ["영어 → 뜻", "뜻 → 영어"], key="t_dir",
                               disabled=(t_mode == "토익식 빈칸"),
                               help="토익식은 항상 영어 단어를 맞히는 방식입니다.")

        pool = filter_words(words, t_cats, t_scope)
        if t_mode == "토익식 빈칸":
            pool = [w for w in pool if make_toeic_question(w, words)]
            st.caption(f"토익식 빈칸 가능 단어: {len(pool)}개")
        else:
            st.caption(f"단어 풀: {len(pool)}개")

        if len(pool) >= 2:
            n_q = st.slider("문제 수", 2, min(len(pool), 50), min(20, len(pool)))
            if st.button("🚀 시험 시작", type="primary", use_container_width=True):
                start_regular_test(pool, n_q, t_mode, t_dir)
                st.rerun()
        else:
            st.warning("조건에 맞는 단어가 2개 미만이에요.")

    elif not st.session_state.test_submitted:
        questions  = st.session_state.test_questions
        mode       = st.session_state.test_mode_val
        direction  = st.session_state.test_dir_val
        toeic_data = st.session_state.test_toeic_data

        st.subheader(f"📝 시험 중 — {mode} · {len(questions)}문제")

        with st.form("test_form"):
            for i, q in enumerate(questions):
                td = toeic_data.get(q["no"])

                if mode == "토익식 빈칸" and td:
                    st.markdown(
                        f"<div class='q-label'><span class='q-num'>Q{i+1}</span> 빈칸에 알맞은 단어를 고르세요.</div>",
                        unsafe_allow_html=True)
                    st.markdown(f"<div class='toeic-sentence'>{td['sentence']}</div>",
                                unsafe_allow_html=True)
                    st.radio("", td["choices"], index=None,
                             label_visibility="collapsed", key=f"test_q_{i}")

                elif mode == "토익식 빈칸" and not td:
                    opts = st.session_state.test_options.get(q["no"], [])
                    st.markdown(
                        f"<div class='q-label'><span class='q-num'>Q{i+1}</span> <code>{q['word']}</code> 의 뜻은?</div>",
                        unsafe_allow_html=True)
                    st.radio("", opts, index=None,
                             label_visibility="collapsed", key=f"test_q_{i}")

                elif mode == "4지선다":
                    opts = st.session_state.test_options.get(q["no"], [])
                    if direction == "영어 → 뜻":
                        st.markdown(
                            f"<div class='q-label'><span class='q-num'>Q{i+1}</span> <code>{q['word']}</code> 의 뜻은?</div>",
                            unsafe_allow_html=True)
                    else:
                        st.markdown(
                            f"<div class='q-label'><span class='q-num'>Q{i+1}</span> <code>{q['meaning']}</code> 에 해당하는 영단어는?</div>",
                            unsafe_allow_html=True)
                    st.radio("", opts, index=None,
                             label_visibility="collapsed", key=f"test_q_{i}")

                else:  # 주관식
                    if direction == "영어 → 뜻":
                        st.markdown(
                            f"<div class='q-label'><span class='q-num'>Q{i+1}</span> <code>{q['word']}</code> 의 뜻은?</div>",
                            unsafe_allow_html=True)
                    else:
                        st.markdown(
                            f"<div class='q-label'><span class='q-num'>Q{i+1}</span> <code>{q['meaning']}</code> 에 해당하는 영단어는?</div>",
                            unsafe_allow_html=True)
                    st.text_input("", placeholder="정답 입력",
                                  label_visibility="collapsed", key=f"test_q_{i}")

                st.divider()

            submitted = st.form_submit_button(
                "✅ 제출하기", type="primary", use_container_width=True)

        if submitted:
            results = []
            for i, q in enumerate(questions):
                user_ans = st.session_state.get(f"test_q_{i}") or ""
                td = toeic_data.get(q["no"])

                if mode == "토익식 빈칸":
                    correct_ans = td["correct"] if td else q["meaning"]
                    is_ok = user_ans == correct_ans
                elif mode == "4지선다":
                    correct_ans = q["meaning"] if direction == "영어 → 뜻" else q["word"]
                    is_ok = user_ans == correct_ans
                else:
                    correct_ans = q["meaning"] if direction == "영어 → 뜻" else q["word"]
                    is_ok = user_ans.lower().strip() == correct_ans.lower().strip()

                if is_ok:
                    mark_correct(q["no"])
                else:
                    mark_wrong(q["no"])

                results.append({"q": q, "user_ans": user_ans,
                                 "correct_ans": correct_ans, "is_correct": is_ok,
                                 "toeic_sentence": td["sentence"] if td else None})

            correct_count = sum(1 for r in results if r["is_correct"])
            score = correct_count / len(results) * 100

            session = {
                "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "total": len(results), "correct": correct_count,
                "score": round(score, 1), "mode": mode, "direction": direction,
            }
            st.session_state.sessions.append(session)
            st.session_state.test_result = {
                "results": results, "score": score, "correct_count": correct_count,
            }
            st.session_state.test_submitted = True
            save_progress()
            st.rerun()

        if st.button("❌ 시험 취소"):
            st.session_state.test_active = False
            st.rerun()

    else:
        res      = st.session_state.test_result
        results  = res["results"]
        score    = res["score"]
        cor_cnt  = res["correct_count"]
        wrong_rs = [r for r in results if not r["is_correct"]]

        st.subheader("📊 시험 결과")
        rc1, rc2, rc3 = st.columns(3)
        rc1.metric("점수",  f"{score:.1f}점")
        rc2.metric("정답",  f"{cor_cnt} / {len(results)}")
        rc3.metric("오답",  f"{len(wrong_rs)}개")

        if   score == 100: st.success("🎉 완벽해요!")
        elif score >= 80:  st.success(f"👍 잘했어요! {score:.0f}점")
        elif score >= 60:  st.warning(f"📚 조금 더 공부해봐요. {score:.0f}점")
        else:              st.error(f"💪 다시 도전해봐요. {score:.0f}점")

        correct_rs = [r for r in results if r["is_correct"]]
        if correct_rs:
            with st.expander(f"🎯 정답 맞춘 단어 ({len(correct_rs)}개)"):
                for r in correct_rs:
                    st.markdown(f"✅ **{r['q']['word']}** — {r['q']['meaning']}")
                buf = make_excel_bytes([r["q"] for r in correct_rs])
                st.download_button(
                    "📥 정답 단어 Excel 저장", data=buf,
                    file_name=f"correct_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_correct")

        if wrong_rs:
            st.subheader("❌ 오답 노트")
            buf = make_excel_bytes([r["q"] for r in wrong_rs])
            st.download_button(
                "📥 오답 Excel 저장", data=buf,
                file_name=f"wrong_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_wrong")
            for r in wrong_rs:
                q = r["q"]
                with st.container(border=True):
                    st.markdown(f"**{q['word']}** — {q['meaning']}")
                    st.caption(f"카테고리: {q['category']}")
                    if r["toeic_sentence"]:
                        st.markdown(f"*{r['toeic_sentence']}*")
                    st.markdown(f"정답: **{r['correct_ans']}** | 내 답: "
                                + (f"~~{r['user_ans']}~~" if r["user_ans"] else "_(미응답)_"))
                    if q["example"]:
                        with st.expander("예문"):
                            st.caption(q["example"])

        btn1, btn2 = st.columns(2)
        with btn1:
            if st.button("🔄 새 시험", use_container_width=True):
                st.session_state.test_active    = False
                st.session_state.test_submitted = False
                st.rerun()
        with btn2:
            if st.button("🔁 오답만 다시 풀기", use_container_width=True,
                         disabled=not wrong_rs):
                pool = [r["q"] for r in wrong_rs]
                start_regular_test(pool, len(pool),
                                   st.session_state.test_mode_val,
                                   st.session_state.test_dir_val)
                st.rerun()


# ============================================================
# TAB 3 — 무한 시험
# ============================================================
with tab_inf:

    if not st.session_state.inf_active:
        st.subheader("♾️ 무한 시험")
        st.caption("단어가 끊임없이 출제됩니다. 맞춘 단어는 🎯 정답 분류에 자동 저장됩니다.")

        inf_cats  = st.multiselect("카테고리", all_categories,
                                   default=all_categories, key="inf_cats")
        ic1, ic2, ic3 = st.columns(3)
        with ic1:
            inf_scope = st.radio("범위", ["전체", "미암기만", "오답만", "정답 맞춘"],
                                 key="inf_scope")
        with ic2:
            inf_mode  = st.radio("유형", ["4지선다", "토익식 빈칸"], key="inf_mode_sel")
        with ic3:
            inf_dir   = st.radio("방향", ["영어 → 뜻", "뜻 → 영어"],
                                 key="inf_dir_sel",
                                 disabled=(inf_mode == "토익식 빈칸"))

        inf_pool = filter_words(words, inf_cats, inf_scope)
        if inf_mode == "토익식 빈칸":
            inf_pool = [w for w in inf_pool if w["example"] and
                        re.search(re.escape(w["word"]), w["example"], re.IGNORECASE)]
        st.caption(f"단어 풀: {len(inf_pool)}개")

        if len(inf_pool) >= 4:
            if st.button("🚀 무한 시험 시작", type="primary", use_container_width=True):
                st.session_state.inf_pool              = inf_pool
                st.session_state.inf_dir               = inf_dir
                st.session_state.inf_mode              = inf_mode
                st.session_state.inf_session_correct   = 0
                st.session_state.inf_session_total     = 0
                st.session_state.inf_session_wrong_nos   = []
                st.session_state.inf_session_correct_nos = []
                st.session_state.inf_active            = True
                pick_next_inf()
                st.rerun()
        else:
            st.warning("단어 풀이 4개 이상 필요해요.")

    else:
        q         = st.session_state.inf_current
        choices   = st.session_state.inf_choices
        correct   = st.session_state.inf_correct_ans
        td        = st.session_state.inf_toeic_data
        mode      = st.session_state.inf_mode
        direction = st.session_state.inf_dir
        s_cor     = st.session_state.inf_session_correct
        s_tot     = st.session_state.inf_session_total
        acc       = s_cor / s_tot * 100 if s_tot > 0 else 0

        hd1, hd2, hd3, hd4 = st.columns(4)
        hd1.metric("✅ 정답", f"{s_cor}개")
        hd2.metric("❌ 오답", f"{s_tot - s_cor}개")
        hd3.metric("정답률",  f"{acc:.1f}%")
        hd4.metric("총 문제", f"{s_tot}개")
        st.divider()

        if mode == "토익식 빈칸" and td:
            st.markdown("**빈칸에 알맞은 단어를 고르세요.**")
            st.markdown(f"<div class='toeic-sentence'>{td['sentence']}</div>",
                        unsafe_allow_html=True)
        else:
            front = q["word"] if direction == "영어 → 뜻" else q["meaning"]
            st.markdown(f"<div class='inf-question'>{front}</div>",
                        unsafe_allow_html=True)
        tts_button(q["word"], key=f"tts_inf_{s_tot}")
        st.caption(f"카테고리: {q['category']}  No.{q['no']}")
        st.write("")

        def handle_inf_answer(selected_choice):
            is_ok = (selected_choice == correct)
            st.session_state.inf_selected  = selected_choice
            st.session_state.inf_correct   = is_ok
            st.session_state.inf_answered  = True
            st.session_state.inf_session_total += 1
            if is_ok:
                st.session_state.inf_session_correct += 1
                if q["no"] not in st.session_state.inf_session_correct_nos:
                    st.session_state.inf_session_correct_nos.append(q["no"])
                mark_correct(q["no"])
            else:
                if q["no"] not in st.session_state.inf_session_wrong_nos:
                    st.session_state.inf_session_wrong_nos.append(q["no"])
                mark_wrong(q["no"])
            save_progress()

        if not st.session_state.inf_answered:
            r1c1, r1c2 = st.columns(2)
            r2c1, r2c2 = st.columns(2)
            btn_slots  = [r1c1, r1c2, r2c1, r2c2]
            for i, (choice, slot) in enumerate(zip(choices, btn_slots)):
                with slot:
                    if st.button(choice, key=f"inf_c_{i}", use_container_width=True):
                        handle_inf_answer(choice)
                        st.rerun()
            st.divider()
            if st.button("⏹️ 시험 종료", use_container_width=True):
                st.session_state.inf_active = False
                st.rerun()

        else:
            selected = st.session_state.inf_selected
            is_ok    = st.session_state.inf_correct

            for choice in choices:
                if choice == correct and choice == selected:
                    st.success(f"✅ {choice}")
                elif choice == correct:
                    st.success(f"✅ {choice}  ← 정답")
                elif choice == selected:
                    st.error(f"❌ {choice}  ← 내 선택")
                else:
                    st.markdown(
                        f"<div style='padding:6px 12px;color:#888'>○ {choice}</div>",
                        unsafe_allow_html=True)

            if is_ok:
                st.success(f"정답입니다! 🎯 **{q['word']}** — {q['meaning']}")
            else:
                st.error(f"오답. 정답: **{correct}**  ({q['word']} — {q['meaning']})")
            tts_button(q["word"], key=f"tts_inf_ans_{s_tot}")

            if q["example"]:
                with st.expander("📖 예문"):
                    st.write(q["example"])

            st.write("")
            nx_col, end_col = st.columns(2)
            with nx_col:
                if st.button("다음 문제 ➡️", type="primary", use_container_width=True):
                    pick_next_inf()
                    st.rerun()
            with end_col:
                if st.button("⏹️ 시험 종료", use_container_width=True):
                    st.session_state.inf_active = False
                    st.rerun()

        col_wrong_panel, col_correct_panel = st.columns(2)
        correct_nos = st.session_state.inf_session_correct_nos
        wrong_nos   = st.session_state.inf_session_wrong_nos

        with col_correct_panel:
            correct_ws = [w for w in words if w["no"] in correct_nos]
            if correct_ws:
                with st.expander(f"🎯 이번 세션 정답 ({len(correct_ws)}개)"):
                    for w in correct_ws:
                        st.markdown(f"- **{w['word']}** — {w['meaning']}")
                    buf = make_excel_bytes(correct_ws)
                    st.download_button("📥 정답 Excel", data=buf,
                                       file_name=f"correct_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="inf_dl_correct")

        with col_wrong_panel:
            wrong_ws = [w for w in words if w["no"] in wrong_nos]
            if wrong_ws:
                with st.expander(f"❌ 이번 세션 오답 ({len(wrong_ws)}개)"):
                    for w in wrong_ws:
                        st.markdown(f"- **{w['word']}** — {w['meaning']}")
                    buf = make_excel_bytes(wrong_ws)
                    st.download_button("📥 오답 Excel", data=buf,
                                       file_name=f"wrong_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key="inf_dl_wrong")


# ============================================================
# TAB 4 — 단어장
# ============================================================
with tab_wordbook:
    st.subheader("📋 단어장")

    wb_cats  = st.multiselect("카테고리", all_categories,
                               default=all_categories, key="wb_cats")
    wc1, wc2 = st.columns(2)
    with wc1:
        wb_scope = st.radio("보기", ["전체", "암기 완료", "미암기", "정답 맞춤", "오답"],
                            horizontal=True, key="wb_scope")
    with wc2:
        search = st.text_input("🔍 단어/뜻 검색", key="wb_search")

    scope_key = {"전체": "전체", "암기 완료": "암기완료",
                 "미암기": "미암기만", "정답 맞춤": "정답 맞춘", "오답": "오답만"}
    wb_pool = filter_words(words, wb_cats, scope_key[wb_scope])
    if wb_scope == "암기 완료":
        wb_pool = [w for w in wb_pool if w["no"] in st.session_state.memorized]
    if search:
        s = search.lower()
        wb_pool = [w for w in wb_pool
                   if s in w["word"].lower() or s in w["meaning"].lower()]

    st.caption(f"{len(wb_pool)}개 표시 중")

    if wb_scope in ("오답", "정답 맞춤") and wb_pool:
        label = "📥 오답 Excel 저장" if wb_scope == "오답" else "📥 정답 맞춤 Excel 저장"
        buf   = make_excel_bytes(wb_pool)
        fname = f"{'wrong' if wb_scope == '오답' else 'correct'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        st.download_button(label, data=buf, file_name=fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    df = pd.DataFrame([{
        "번호": w["no"], "단어": w["word"], "뜻": w["meaning"],
        "예문": (w["example"][:50] + "…") if len(w["example"]) > 50 else w["example"],
        "카테고리": w["category"],
        "🎯": "🎯" if w["no"] in st.session_state.correct_words else "",
        "✅": "✅" if w["no"] in st.session_state.memorized else "",
        "❌": "❌" if w["no"] in st.session_state.wrong else "",
    } for w in wb_pool])
    st.dataframe(df, use_container_width=True, hide_index=True)


# ============================================================
# TAB 5 — 단어 관리 (추가 / 수정 / 삭제)
# ============================================================
with tab_mgmt:
    st.subheader("✏️ 단어 관리")

    mgmt_action = st.radio("작업 선택", ["➕ 단어 추가", "✏️ 단어 수정", "🗑️ 단어 삭제"],
                           horizontal=True, key="mgmt_action")

    # ── 단어 추가 ────────────────────────────────────────────
    if mgmt_action == "➕ 단어 추가":
        st.markdown("#### 새 단어 추가")
        with st.form("add_word_form", clear_on_submit=True):
            new_word     = st.text_input("영어 단어 *", placeholder="예: inevitable")
            new_meaning  = st.text_input("뜻 *", placeholder="예: 불가피한, 필연적인")
            new_example  = st.text_area("예문 (선택)", placeholder="예: Change is inevitable.", height=80)
            new_category = st.selectbox("카테고리", all_categories + ["직접 입력"])
            if new_category == "직접 입력":
                new_category = st.text_input("새 카테고리 이름")

            submitted_add = st.form_submit_button("➕ 추가하기", type="primary",
                                                   use_container_width=True)

        if submitted_add:
            if not new_word.strip() or not new_meaning.strip():
                st.error("영어 단어와 뜻은 필수입니다.")
            else:
                new_no = max((w["no"] for w in words if w["no"]), default=0) + 1
                new_entry = {
                    "no":       new_no,
                    "word":     new_word.strip(),
                    "meaning":  new_meaning.strip(),
                    "example":  new_example.strip(),
                    "category": new_category.strip() or "기타",
                    "_extra":   [None, None, None, None],
                }
                updated_words = words + [new_entry]
                save_words_to_excel(updated_words)
                st.success(f"✅ **{new_word.strip()}** 단어가 추가되었습니다!")
                st.rerun()

    # ── 단어 수정 ────────────────────────────────────────────
    elif mgmt_action == "✏️ 단어 수정":
        st.markdown("#### 단어 수정")
        edit_search = st.text_input("🔍 수정할 단어 검색", key="edit_search",
                                    placeholder="영어 단어 또는 뜻을 입력하세요")
        if edit_search:
            s = edit_search.lower()
            found = [w for w in words
                     if s in w["word"].lower() or s in w["meaning"].lower()]
            if not found:
                st.info("검색 결과가 없습니다.")
            else:
                for w in found[:20]:
                    col_info, col_btn = st.columns([4, 1])
                    with col_info:
                        st.markdown(f"**{w['word']}** — {w['meaning']}  "
                                    f"<small>({w['category']})</small>",
                                    unsafe_allow_html=True)
                    with col_btn:
                        if st.button("수정", key=f"edit_btn_{w['no']}",
                                     use_container_width=True):
                            st.session_state.mgmt_edit_no = w["no"]

        if st.session_state.mgmt_edit_no is not None:
            target = next((w for w in words
                           if w["no"] == st.session_state.mgmt_edit_no), None)
            if target:
                st.divider()
                st.markdown(f"#### ✏️ 수정 중: **{target['word']}**")
                with st.form("edit_word_form"):
                    ed_word     = st.text_input("영어 단어", value=target["word"])
                    ed_meaning  = st.text_input("뜻",       value=target["meaning"])
                    ed_example  = st.text_area("예문",      value=target["example"], height=80)
                    cat_options = all_categories if target["category"] in all_categories \
                                  else all_categories + [target["category"]]
                    ed_category = st.selectbox("카테고리", cat_options,
                                               index=cat_options.index(target["category"]))
                    col_save, col_cancel = st.columns(2)
                    with col_save:
                        save_edit = st.form_submit_button("💾 저장", type="primary",
                                                           use_container_width=True)
                    with col_cancel:
                        cancel_edit = st.form_submit_button("취소", use_container_width=True)

                if save_edit:
                    if not ed_word.strip() or not ed_meaning.strip():
                        st.error("영어 단어와 뜻은 필수입니다.")
                    else:
                        updated_words = []
                        for w in words:
                            if w["no"] == st.session_state.mgmt_edit_no:
                                updated_words.append({
                                    **w,
                                    "word":     ed_word.strip(),
                                    "meaning":  ed_meaning.strip(),
                                    "example":  ed_example.strip(),
                                    "category": ed_category,
                                })
                            else:
                                updated_words.append(w)
                        save_words_to_excel(updated_words)
                        st.session_state.mgmt_edit_no = None
                        st.success("✅ 단어가 수정되었습니다!")
                        st.rerun()

                if cancel_edit:
                    st.session_state.mgmt_edit_no = None
                    st.rerun()

    # ── 단어 삭제 ────────────────────────────────────────────
    else:
        st.markdown("#### 단어 삭제")
        st.warning("삭제한 단어는 복구할 수 없습니다. 신중하게 사용하세요.")
        del_search = st.text_input("🔍 삭제할 단어 검색", key="del_search",
                                   placeholder="영어 단어 또는 뜻을 입력하세요")
        if del_search:
            s = del_search.lower()
            found = [w for w in words
                     if s in w["word"].lower() or s in w["meaning"].lower()]
            if not found:
                st.info("검색 결과가 없습니다.")
            else:
                for w in found[:20]:
                    col_info, col_btn = st.columns([4, 1])
                    with col_info:
                        st.markdown(f"**{w['word']}** — {w['meaning']}  "
                                    f"<small>({w['category']})</small>",
                                    unsafe_allow_html=True)
                    with col_btn:
                        if st.button("삭제", key=f"del_btn_{w['no']}",
                                     use_container_width=True, type="primary"):
                            st.session_state.mgmt_confirm_delete = w["no"]

        if st.session_state.mgmt_confirm_delete is not None:
            target = next((w for w in words
                           if w["no"] == st.session_state.mgmt_confirm_delete), None)
            if target:
                st.divider()
                st.error(f"**{target['word']}** ({target['meaning']}) 을(를) 삭제하시겠습니까?")
                conf1, conf2 = st.columns(2)
                with conf1:
                    if st.button("🗑️ 삭제 확인", type="primary", use_container_width=True):
                        updated_words = [w for w in words
                                         if w["no"] != st.session_state.mgmt_confirm_delete]
                        save_words_to_excel(updated_words)
                        # 관련 진도 데이터도 제거
                        no = st.session_state.mgmt_confirm_delete
                        st.session_state.memorized.discard(no)
                        st.session_state.wrong.discard(no)
                        st.session_state.correct_words.discard(no)
                        save_progress()
                        st.session_state.mgmt_confirm_delete = None
                        st.success("🗑️ 단어가 삭제되었습니다.")
                        st.rerun()
                with conf2:
                    if st.button("취소", use_container_width=True):
                        st.session_state.mgmt_confirm_delete = None
                        st.rerun()


# ============================================================
# TAB 6 — 통계
# ============================================================
with tab_stats:
    st.subheader("📊 학습 통계")

    sc1, sc2, sc3, sc4, sc5 = st.columns(5)
    sc1.metric("전체",        f"{total}개")
    sc2.metric("암기 완료",   f"{mem_count}개",  delta=f"{mem_count/total*100:.1f}%")
    sc3.metric("🎯 정답 맞춤", f"{cor_count}개",  delta=f"{cor_count/total*100:.1f}%")
    sc4.metric("❌ 오답",     f"{len(st.session_state.wrong)}개")
    sc5.metric("미분류",
               f"{total - len(st.session_state.memorized | st.session_state.correct_words | st.session_state.wrong)}개")

    dl1, dl2 = st.columns(2)
    with dl1:
        if st.session_state.correct_words:
            cw = [w for w in words if w["no"] in st.session_state.correct_words]
            buf = make_excel_bytes(cw)
            st.download_button(
                f"📥 전체 정답 단어 Excel ({len(cw)}개)", data=buf,
                file_name=f"correct_all_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with dl2:
        if st.session_state.wrong:
            ww = [w for w in words if w["no"] in st.session_state.wrong]
            buf = make_excel_bytes(ww)
            st.download_button(
                f"📥 전체 오답 단어 Excel ({len(ww)}개)", data=buf,
                file_name=f"wrong_all_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.divider()
    st.subheader("카테고리별 현황")
    cat_rows = []
    for cat in all_categories:
        cw = [w for w in words if w["category"] == cat]
        cat_rows.append({
            "카테고리": cat,
            "전체": len(cw),
            "✅ 암기": len([w for w in cw if w["no"] in st.session_state.memorized]),
            "🎯 정답": len([w for w in cw if w["no"] in st.session_state.correct_words]),
            "❌ 오답": len([w for w in cw if w["no"] in st.session_state.wrong]),
        })
        cat_rows[-1]["진도율"] = (
            f"{(cat_rows[-1]['✅ 암기'] + cat_rows[-1]['🎯 정답']) / len(cw) * 100:.1f}%"
            if cw else "0%"
        )
    st.dataframe(pd.DataFrame(cat_rows), use_container_width=True, hide_index=True)

    st.subheader("시험 기록")
    sessions = st.session_state.sessions
    if sessions:
        df_s = pd.DataFrame(sessions[-20:][::-1]).rename(columns={
            "date": "날짜", "total": "문제수", "correct": "정답",
            "score": "점수", "mode": "유형", "direction": "방향",
        })
        st.dataframe(df_s, use_container_width=True, hide_index=True)
        avg  = sum(s["score"] for s in sessions) / len(sessions)
        best = max(s["score"] for s in sessions)
        sh1, sh2, sh3 = st.columns(3)
        sh1.metric("총 시험 횟수", f"{len(sessions)}회")
        sh2.metric("평균 점수",    f"{avg:.1f}점")
        sh3.metric("최고 점수",    f"{best:.1f}점")
    else:
        st.info("아직 시험 기록이 없어요.")

    st.divider()
    rst1, rst2, rst3 = st.columns(3)
    with rst1:
        if st.button("🗑️ 암기 체크 초기화"):
            st.session_state.memorized = set()
            save_progress(); st.rerun()
    with rst2:
        if st.button("🗑️ 오답 기록 초기화"):
            st.session_state.wrong = set()
            save_progress(); st.rerun()
    with rst3:
        if st.button("🗑️ 전체 진도 초기화"):
            st.session_state.memorized     = set()
            st.session_state.wrong         = set()
            st.session_state.correct_words = set()
            st.session_state.sessions      = []
            save_progress(); st.rerun()
